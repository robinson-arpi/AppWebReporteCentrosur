import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import streamlit as st
import io
from datetime import datetime, timedelta


# Variables globales para almacenar nombres de las hojas del archivo Excel
sheet_names = []

# Estilos de formato para el reporte en Excel
bold_font_title = Font(size=14)  # Fuente en negrita con tamaño de 14
bold_font = Font(bold=True)  # Fuente en negrita
highlight = PatternFill("solid", fgColor="FFFF00")  # Resaltado en amarillo
header_fill = PatternFill("solid", fgColor="dbf3d3")  # Relleno verde claro para cabeceras
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))  # Borde delgado en celdas

# Verifica si una fila está completamente vacía
def empty_row(fila):
    return fila.isnull().all()

# Función para leer el archivo Excel y concatenar todas las hojas
def read_excel(input_file):
    try:
        # Leer todas las hojas del archivo Excel
        df = pd.read_excel(input_file, sheet_name=None)  # Lee todas las hojas en un diccionario
        valid_rows = []  # Lista para almacenar filas válidas (no vacías)

        # Iterar sobre cada hoja de Excel
        for sheet_name, data in df.items():
            for index, fila in data.iterrows():
                if empty_row(fila):  # Si la fila está vacía, detiene la iteración
                    break
                valid_rows.append(fila)  # Agrega la fila si es válida

        # Crear un DataFrame que contiene todas las filas válidas
        valid_df = pd.DataFrame(valid_rows)
        
        # Obtener los nombres de las hojas y limpiar los nombres de las columnas
        sheet_names = list(df.keys())
        valid_df.columns = valid_df.columns.str.strip().str.replace('\n', ' ', regex=True).str.replace('_', ' ')  # Limpieza de columnas
        
        return valid_df, sheet_names

    except Exception as e:
        st.write(f"Error en read_excel: {e}")  # Manejo de excepciones con detalle del error
        return None, None

# Se presentó un problema respecto a la selección de sectores
# En caso de haber problema  se  reporta en la interfaz
def check_sectors(df):
    try:
        df['SECTORES'] = df['SECTORES'].str.replace('\n', ' ').str.replace('\r', ' ').str.strip()

        # Clave para verificar que los sectores sena correctos ['CANTON', 'ZONA', 'NUMERO CLIENTES']
        groupings = df.groupby(['CANTON', 'ZONA', 'NUMERO CLIENTES'])
        corrections = {}
        rows_with_error = []

        for (canton, zone, num_clientes), grupo in groupings:
            sectores = grupo['SECTORES'].tolist()
            if len(set(sectores)) > 1:
                sector_mayor = max(sectores, key=len)
                nuevo_sector = sector_mayor
                corrections[(canton, zone, num_clientes)] = nuevo_sector
                rows_with_error.extend(grupo.index.tolist())

        for (canton, zone, num_clientes), nuevo_sector in corrections.items():
            df.loc[(df['CANTON'] == canton) & (df['ZONA'] == zone) & (df['NUMERO CLIENTES'] == num_clientes), 'SECTORES'] = nuevo_sector

        return df
    except Exception as e:
        st.write("Error en check_sectors:" + str(e))
        return None


def combine_hours(group):
    try:
        ordered_hours = group[['HORA INICIO', 'HORA FINAL']].sort_values(by='HORA INICIO')
        return ' '.join([f"{row['HORA INICIO']}-{row['HORA FINAL']}" for index, row in ordered_hours.iterrows()])
    except Exception as e:
        st.write("Error en combine_hours:" + str(e))
        return None
    


def create_worksheet(wb, df_agrupado, day, start_column=3):
    # Convertir el número de la columna de inicio en una letra
    start_col_letter = get_column_letter(start_column)
    
    ws = wb.create_sheet(title=f"Dia {day.replace('/', '-')}")

    # Ajustar las posiciones de acuerdo al valor de start_column
    ws[f"{start_col_letter}2"] = "FORMATO DE DESCONEXIONES DIARIAS"
    ws.merge_cells(f'{start_col_letter}2:{get_column_letter(start_column + 2)}2')
    ws[f"{start_col_letter}3"] = "EMPRESA:"
    ws[f"{get_column_letter(start_column + 1)}3"] = "CENTROSUR"
    ws.merge_cells(f'{get_column_letter(start_column + 1)}3:{get_column_letter(start_column + 2)}3')
    ws[f"{start_col_letter}4"] = "FECHA:"
    ws[f"{get_column_letter(start_column + 1)}4"] = day.replace('/', '-')
    ws.merge_cells(f'{get_column_letter(start_column + 1)}4:{get_column_letter(start_column + 2)}4')

    # Aplicar estilos
    ws[f"{start_col_letter}2"].font = bold_font_title
    ws[f"{start_col_letter}3"].font = bold_font_title
    ws[f"{start_col_letter}4"].font = bold_font_title
    ws[f"{get_column_letter(start_column + 1)}3"].font = bold_font_title
    ws[f"{get_column_letter(start_column + 1)}4"].font = bold_font_title

    ws[f"{start_col_letter}2"].border = thin_border
    ws[f"{start_col_letter}3"].border = thin_border
    ws[f"{start_col_letter}4"].border = thin_border
    ws[f"{get_column_letter(start_column+2)}2"].border = thin_border
    ws[f"{get_column_letter(start_column + 1)}3"].border = thin_border
    ws[f"{get_column_letter(start_column + 2)}3"].border = thin_border
    ws[f"{get_column_letter(start_column + 1)}4"].border = thin_border
    ws[f"{get_column_letter(start_column + 2)}4"].border = thin_border
    
    ws[f"{get_column_letter(start_column + 1)}3"].fill = header_fill
    ws[f"{get_column_letter(start_column + 1)}4"].fill = header_fill
    


    row = 6
    contador = 1
    df_periodos = df_agrupado.groupby('PERIODO')

    # Iterar sobre los periodos agrupados
    for periodo, datos in df_periodos:
        # Ajustar las posiciones de acuerdo a start_column
        horas = calcular_horas(datos["PERIODO"].iloc[0])
        ws[f"{start_col_letter}{row}"] = f"BLOQUE {contador}"
        ws[f"{get_column_letter(start_column + 1)}{row}"] = "SUBESTACIÓN"
        ws[f"{get_column_letter(start_column + 2)}{row}"] = "PRIMARIOS A DESCONECTAR"
        ws[f"{get_column_letter(start_column + 3)}{row}"] = '# CLIENTES'
        ws[f"{get_column_letter(start_column + 6)}{row}"] = 'DEMANDA PROMEDIO DE LOS PERIODOS (MWh)'
        ws[f"{get_column_letter(start_column + 9)}{row}"] = "PROVINCIA"
        ws[f"{get_column_letter(start_column + 10)}{row}"] = "CANTON"
        ws[f"{get_column_letter(start_column + 11)}{row}"] = "SECTORES"


        ws[f"{get_column_letter(start_column + 3)}{row+1}"] = 'RESIDENCIAL'
        ws[f"{get_column_letter(start_column + 4)}{row+1}"] = 'INDUSTRIAL'
        ws[f"{get_column_letter(start_column + 5)}{row+1}"] = 'COMERCIAL'
        ws[f"{get_column_letter(start_column + 6)}{row+1}"] = 'RESIDENCIAL'
        ws[f"{get_column_letter(start_column + 7)}{row+1}"] = 'INDUSTRIAL'
        ws[f"{get_column_letter(start_column + 8)}{row+1}"] = 'COMERCIAL'

        # Aplicar formato a las celdas del encabezado
# Aplicar estilos a la fila actual y la siguiente fila en un solo bucle
        for col in range(start_column, start_column + 12):
            for r in range(row, row + 2):  # Aplica los estilos a la fila actual (row) y la siguiente (row + 1)
                cell = ws.cell(row=r, column=col)
                cell.font = bold_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws.merge_cells(f"{start_col_letter}{row}:{start_col_letter}{row+1}")
        ws.merge_cells(f"{get_column_letter(start_column + 1)}{row}:{get_column_letter(start_column + 1)}{row+1}")
        ws.merge_cells(f"{get_column_letter(start_column + 2)}{row}:{get_column_letter(start_column + 2)}{row+1}")        
        ws.merge_cells(f"{get_column_letter(start_column + 9)}{row}:{get_column_letter(start_column + 9)}{row+1}")        
        ws.merge_cells(f"{get_column_letter(start_column + 10)}{row}:{get_column_letter(start_column + 10)}{row+1}")        
        ws.merge_cells(f"{get_column_letter(start_column + 11)}{row}:{get_column_letter(start_column + 11)}{row+1}")        

        ws.merge_cells(f"{get_column_letter(start_column + 3)}{row}:{get_column_letter(start_column + 5)}{row}")        
        ws.merge_cells(f"{get_column_letter(start_column + 6)}{row}:{get_column_letter(start_column + 8)}{row}")        

        row +=1

        if isinstance(datos, pd.Series):
            datos = datos.to_frame().T
        merge_start = row + 1
        for _, fila in datos.iterrows():
            carga_est = float(fila["CARGA EST MW"])
            # Insertar fila a partir de start_column
            ws.append([None] * (start_column - 1) + list(fila))
            # Recuperar la fila actual donde estamos añadiendo datos
            current_row = ws.max_row
            # Recuperar el valor actual en la columna start_column + 6
            current_value = ws[f"{get_column_letter(start_column + 6)}{current_row}"].value
            ws[f"{get_column_letter(start_column + 6)}{current_row}"] = f"={current_value}*{horas}*{carga_est}"
            current_value = ws[f"{get_column_letter(start_column + 7)}{current_row}"].value
            ws[f"{get_column_letter(start_column + 7)}{current_row}"] = f"={current_value}*{horas}*{carga_est}"
            current_value = ws[f"{get_column_letter(start_column + 8)}{current_row}"].value
            ws[f"{get_column_letter(start_column + 8)}{current_row}"] = f"={current_value}*{horas}*{carga_est}"

            row += 1
        merge_end = row

        # Aplicar borde a las celdas del rango de datos
        for r in range(row - len(datos), row + 3):
            for c in range(start_column, start_column + len(fila)):
                cell = ws.cell(row=r, column=c)
                cell.border = thin_border

        contador += 1
        row += 4
        #MErge periodos
        ws.merge_cells(f'{start_col_letter}{merge_start}:{start_col_letter}{merge_end+2}')

        # Centrar el contenido después del merge
        merged_cell = ws[f"{start_col_letter}{merge_start}"]
        merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)



        # Usar la fórmula de Excel para sumar el rango
        ws[f"{get_column_letter(start_column + 1)}{merge_end + 1}"] = "TOTALES PARCIALES:"
        ws[f"{get_column_letter(start_column + 1)}{merge_end + 2}"] = "TOTAL:"

        
        ws[f"{get_column_letter(start_column + 3)}{merge_end + 1}"] = f"=SUM({get_column_letter(start_column + 3)}{merge_start}:{get_column_letter(start_column + 3)}{merge_end})"
        ws[f"{get_column_letter(start_column + 4)}{merge_end + 1}"] = f"=SUM({get_column_letter(start_column + 4)}{merge_start}:{get_column_letter(start_column + 4)}{merge_end})"
        ws[f"{get_column_letter(start_column + 5)}{merge_end + 1}"] = f"=SUM({get_column_letter(start_column + 5)}{merge_start}:{get_column_letter(start_column + 5)}{merge_end})"
        ws[f"{get_column_letter(start_column + 6)}{merge_end + 1}"] = f"=SUM({get_column_letter(start_column + 6)}{merge_start}:{get_column_letter(start_column + 6)}{merge_end})"
        ws[f"{get_column_letter(start_column + 7)}{merge_end + 1}"] = f"=SUM({get_column_letter(start_column + 7)}{merge_start}:{get_column_letter(start_column + 7)}{merge_end})"
        ws[f"{get_column_letter(start_column + 8)}{merge_end + 1}"] = f"=SUM({get_column_letter(start_column + 8)}{merge_start}:{get_column_letter(start_column + 8)}{merge_end})"

        # Sumar las tres primeras columnas (start_column + 3, start_column + 4, start_column + 5) usando SUM
        ws[f"{get_column_letter(start_column + 3)}{merge_end + 2}"] = (
            f"=SUM({get_column_letter(start_column + 3)}{merge_end + 1}, " 
            f"{get_column_letter(start_column + 4)}{merge_end + 1}, "
            f"{get_column_letter(start_column + 5)}{merge_end + 1})"
        )

        # Sumar las siguientes tres columnas (start_column + 6, start_column + 7, start_column + 8) usando SUM
        ws[f"{get_column_letter(start_column + 6)}{merge_end + 2}"] = (
            f"=SUM({get_column_letter(start_column + 6)}{merge_end + 1}, "
            f"{get_column_letter(start_column + 7)}{merge_end + 1}, "
            f"{get_column_letter(start_column + 8)}{merge_end + 1})"
        )


        #Merges para calculos
        ws.merge_cells(f"{get_column_letter(start_column + 1)}{merge_end + 1}:{get_column_letter(start_column + 2)}{merge_end + 1}")
        ws.merge_cells(f"{get_column_letter(start_column + 1)}{merge_end + 2}:{get_column_letter(start_column + 2)}{merge_end + 2}")

        ws.merge_cells(f"{get_column_letter(start_column + 3)}{merge_end + 2}:{get_column_letter(start_column + 5)}{merge_end + 2}")
        ws.merge_cells(f"{get_column_letter(start_column + 6)}{merge_end + 2}:{get_column_letter(start_column + 8)}{merge_end + 2}")
        
        ws.merge_cells(f"{get_column_letter(start_column + 9)}{merge_end + 1}:{get_column_letter(start_column + 11)}{merge_end + 2}")

        #Formato
        ws[f"{get_column_letter(start_column + 1)}{merge_end + 1}"].font = bold_font
        ws[f"{get_column_letter(start_column + 1)}{merge_end + 2}"].font = bold_font
        
        for r in range(merge_start, merge_end+3):
            ws.cell(row=r, column=start_column + 3).number_format = '0.0'
            ws.cell(row=r, column=start_column + 4).number_format = '0.0'
            ws.cell(row=r, column=start_column + 5).number_format = '0.0'
            ws.cell(row=r, column=start_column + 6).number_format = '0.00'
            ws.cell(row=r, column=start_column + 7).number_format = '0.00'
            ws.cell(row=r, column=start_column + 8).number_format = '0.00'
            ws.cell(row=r, column=start_column + 11).alignment = Alignment(vertical='center', wrap_text=True)
            # Alinear todas las columnas a la izquierda
            ws.cell(row=r, column=start_column + 1).alignment = Alignment(horizontal='left',vertical='top')
            ws.cell(row=r, column=start_column + 3).alignment = Alignment(horizontal='left',vertical='top')
            ws.cell(row=r, column=start_column + 4).alignment = Alignment(horizontal='left',vertical='top')
            ws.cell(row=r, column=start_column + 5).alignment = Alignment(horizontal='left',vertical='top')
            ws.cell(row=r, column=start_column + 6).alignment = Alignment(horizontal='left',vertical='top')
            ws.cell(row=r, column=start_column + 7).alignment = Alignment(horizontal='left',vertical='top')
            ws.cell(row=r, column=start_column + 8).alignment = Alignment(horizontal='left',vertical='top')

            ws.cell(row=r, column=start_column + 2).alignment = Alignment(vertical='top')
            ws.cell(row=r, column=start_column + 9).alignment = Alignment(vertical='top')
            ws.cell(row=r, column=start_column + 10).alignment = Alignment(vertical='top')





    # Aplicar formato a los números en la columna de carga estimada
    # for r in range(2, row):
    #     ws.cell(row=r, column=start_column + 3).number_format = '0.0'
    #     ws.cell(row=r, column=start_column + 4).number_format = '0.0'
    #     ws.cell(row=r, column=start_column + 5).number_format = '0.0'
    #     ws.cell(row=r, column=start_column + 6).number_format = '0.00'
    #     ws.cell(row=r, column=start_column + 7).number_format = '0.00'
    #     ws.cell(row=r, column=start_column + 8).number_format = '0.00'
    #     ws.cell(row=r, column=start_column + 11).alignment = Alignment(vertical='center', wrap_text=True)
    #     # Alinear todas las columnas a la izquierda
    #     ws.cell(row=r, column=start_column + 1).alignment = Alignment(horizontal='left')
    #     ws.cell(row=r, column=start_column + 3).alignment = Alignment(horizontal='left')
    #     ws.cell(row=r, column=start_column + 4).alignment = Alignment(horizontal='left')
    #     ws.cell(row=r, column=start_column + 5).alignment = Alignment(horizontal='left')
    #     ws.cell(row=r, column=start_column + 6).alignment = Alignment(horizontal='left')
    #     ws.cell(row=r, column=start_column + 7).alignment = Alignment(horizontal='left')
    #     ws.cell(row=r, column=start_column + 8).alignment = Alignment(horizontal='left')



    # Ajustar ancho de las columnas dinámicamente a partir de start_column
    ancho_columnas = {
        0: 20,  # Columna A (20)
        1: 16,  # Columna B (16)
        2: 25,  # Columna C (25)
        3: 15,  # Columna D (15)
        4: 15,  # Columna E (15)
        5: 15,  # Columna F (15)
        6: 15,  # Columna G (20)
        7: 15,  # Columna H (15)
        8: 15,   # Columna I (15)
        9: 15,   # Columna J (15)
        10: 15,   # Columna K (15)
        11: 180,   # Columna L (15)


    }

    for idx, ancho in ancho_columnas.items():
        col_letter = get_column_letter(start_column + idx)
        ws.column_dimensions[col_letter].width = ancho

def procesar_datos_por_dia(df):
    # Convertir la columna DIA a datetime
    df['DIA'] = pd.to_datetime(df['DIA'], format='%d/%m/%Y', errors='coerce')
    # Agrupar por día con formato seguro para el nombre de la hoja
    df_por_dia = {day.strftime('%Y-%m-%d'): datos for day, datos in df.groupby(df['DIA'])}
    return df_por_dia


def calcular_horas(periodos):
    total_horas = 0
    periodos = periodos.split()  # Divide los diferentes bloques de tiempo

    for periodo in periodos:
        inicio, fin = periodo.split('-')  # Separa las horas de inicio y fin
        fmt = '%H:%M:%S'  # Formato de las horas
        t_inicio = datetime.strptime(inicio, fmt)
        t_fin = datetime.strptime(fin, fmt)

        # Si la hora de fin es menor que la de inicio, sumamos 1 día (24 horas)
        if t_fin < t_inicio:
            t_fin += timedelta(days=1)

        # Calcula la diferencia en horas y añade al total
        horas = (t_fin - t_inicio).total_seconds() / 3600
        total_horas += horas

    return total_horas

# Info page
st.set_page_config(
    page_title="Reporte",
    page_icon='images/icono-centrosur.ico',
    layout="centered",
    initial_sidebar_state="expanded",  # Para que la barra lateral esté siempre expandida
)
# Sidebar para instrucciones
logo_url = 'images/logo-centrosur.png'
st.sidebar.image(logo_url)
st.sidebar.header("Instrucciones")
st.sidebar.write("""
Por favor, suba un archivo Excel sin ninguna edición, debería contar con al menos las siguientes cabeceras:
- HORA INICIO
- HORA FINAL
- DIA
- BLOQUE
- SUBESTACIÓN
- PRIMARIOS A DESCONECTAR
- EQUIPO ABRIR
- EQUIPO TRANSF
- CARGA EST MW
- PROVINCIA
- CANTON
- ZONA
- SECTORES
- Prevalencia del Alimentador CTipo de Cliente)
- NUMERO CLIENTES

El programa procesará los datos, separará por días y generará un reporte para descargar.
""")

# Streamlit Interface
st.title("Reporte de cortes de energía Centrosur")

uploaded_file = st.file_uploader("Elige un archivo Excel", type="xlsx")

st.divider()
if uploaded_file:
    try:
        df, sheet_names = read_excel(uploaded_file)
        
        # Crear dos columnas
        col1, col2 = st.columns(2)
        with col1:
            st.write("Hojas encontradas:", sheet_names)

        # Definir las columnas requeridas
        required_columns = ['SECTORES', 'SUBESTACIÓN', 'CARGA EST MW', 'HORA INICIO', 
                            'HORA FINAL', 'PROVINCIA', 'PRIMARIOS A DESCONECTAR', 
                            'CANTON', 'Prevalencia del Alimentador CTipo de Cliente)', 
                            'NUMERO CLIENTES', 'DIA', 'ZONA', 'CLIENTES RESIDENCIALES','Aporte Residencial','CLIENTES COMERCIALES' ,'Aporte Comercial','CLIENTES INDUSTRIALES', 'Aporte Industrial']

        # Verificar si las columnas requeridas están en el DataFrame
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise KeyError(f"Faltan las siguientes columnas: {', '.join(missing_columns)}")
        
        # Seleccionar las columnas requeridas del DataFrame
        df_seleccionado = df[required_columns]
        
        df_seleccionado = check_sectors(df_seleccionado)

        df_seleccionado.loc[:, 'DIA'] = pd.to_datetime(df_seleccionado['DIA'], format='%d/%m/%Y')
        # Procesar los datos por día
        df_por_dia = procesar_datos_por_dia(df_seleccionado)

        # Crear un nuevo libro de trabajo
        wb = Workbook()

        for day, df_agrupado in df_por_dia.items():
            df_agrupado = df_agrupado.groupby('PRIMARIOS A DESCONECTAR').agg({
                'HORA INICIO': lambda x: list(x),
                'HORA FINAL': lambda x: list(x),
                'SUBESTACIÓN': 'first',
                'CARGA EST MW': 'first',
                'PROVINCIA': 'first',
                'CANTON': 'first',
                'SECTORES': 'first',
                'Prevalencia del Alimentador CTipo de Cliente)': 'first',
                'NUMERO CLIENTES': 'sum',
                'ZONA': 'first',
                'NUMERO CLIENTES': 'first',
                'CLIENTES RESIDENCIALES':'first',
                'CLIENTES INDUSTRIALES': 'first',
                'CLIENTES COMERCIALES': 'first',
                'Aporte Residencial': 'mean',
                'Aporte Industrial': 'mean',
                'Aporte Comercial': 'mean'

            }).reset_index()

            df_agrupado['PERIODO'] = df_agrupado.apply(lambda row: combine_hours(pd.DataFrame({'HORA INICIO': row['HORA INICIO'], 'HORA FINAL': row['HORA FINAL']})), axis=1)
            df_agrupado = df_agrupado.sort_values(by='PERIODO')
            df_agrupado = df_agrupado[['PERIODO', 'SUBESTACIÓN', 'PRIMARIOS A DESCONECTAR','CLIENTES RESIDENCIALES','CLIENTES INDUSTRIALES','CLIENTES COMERCIALES','Aporte Residencial',  'Aporte Industrial','Aporte Comercial', 'PROVINCIA', 'CANTON', 'SECTORES', 'CARGA EST MW']]

            # Crear una hoja por cada día
            create_worksheet(wb, df_agrupado, day)

        # Eliminar la hoja por defecto llamada "Sheet"
        if "Sheet" in wb.sheetnames:
            std_sheet = wb["Sheet"]
            wb.remove(std_sheet)

        # Guardar el archivo en un objeto BytesIO
        output_file = io.BytesIO()
        wb.save(output_file)
        output_file.seek(0)

        with col2:
            st.write("Reporte generado:")
            
            # Botón de descarga
            st.download_button(
                label="Descargar archivo Excel",
                data=output_file,
                file_name='Formato MEM.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
    except KeyError as ke:
        st.write("Error de índice: Asegúrate de que el DataFrame tiene las columnas requeridas.")
        st.write("Detalles del error: " + str(ke))
        
    except Exception as e:
        st.write("Error en main: " + str(e))
