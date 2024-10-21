from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import pandas as pd
from datetime import datetime, timedelta

# Estilos de formato para el reporte en Excel
bold_font_title = Font(size=14)  # Fuente en negrita con tamaño de 14
bold_font = Font(bold=True)  # Fuente en negrita
highlight = PatternFill("solid", fgColor="FFFF00")  # Resaltado en amarillo
header_fill = PatternFill("solid", fgColor="dbf3d3")  # Relleno verde claro para cabeceras
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))  # Borde delgado en celdas


def calculate_hours(periods):
    total_hours = 0
    periods = periods.split()  # Divide los diferentes bloques de tiempo

    for period in periods:
        start, end = period.split('-')  # Separa las horas de inicio y fin
        fmt = '%H:%M:%S'  # Formato de las horas
        t_start = datetime.strptime(start, fmt)
        t_end = datetime.strptime(end, fmt)

        # Si la hora de fin es menor que la de inicio, sumamos 1 día (24 horas)
        if t_end < t_start:
            t_end += timedelta(days=1)

        # Calcula la diferencia en horas y añade al total
        hours = (t_end - t_start).total_seconds() / 3600
        total_hours += hours

    return total_hours

def create_worksheet(wb, df_agrupado, day, start_column=3):
    # Convertir el número de la columna de inicio en una letra
    start_col_letter = get_column_letter(start_column)
    
    ws = wb.create_sheet(title=f"Dia {day.replace('/', '-')}")

    # Ajustar las posiciones de acuerdo al valor de start_column
    ws[f"{start_col_letter}2"] = "FORMATO DE DESCONEXIONES DIARIAS"
    ws.merge_cells(f'{start_col_letter}2:{get_column_letter(start_column + 2)}2')
    ws[f"{start_col_letter}3"] = "EMPRESA:"
    ws[f"{get_column_letter(start_column + 1)}3"] = "CENTROSUR"
    ws.merge_cells(f'{get_column_letter(start_column + 1)}3:{get_column_letter(start_column + 2)}3')
    ws[f"{start_col_letter}4"] = "FECHA:"
    ws[f"{get_column_letter(start_column + 1)}4"] = day.replace('/', '-')
    ws.merge_cells(f'{get_column_letter(start_column + 1)}4:{get_column_letter(start_column + 2)}4')

    # Aplicar estilos
    ws[f"{start_col_letter}2"].font = bold_font_title
    ws[f"{start_col_letter}3"].font = bold_font_title
    ws[f"{start_col_letter}4"].font = bold_font_title
    ws[f"{get_column_letter(start_column + 1)}3"].font = bold_font_title
    ws[f"{get_column_letter(start_column + 1)}4"].font = bold_font_title

    ws[f"{start_col_letter}2"].border = thin_border
    ws[f"{start_col_letter}3"].border = thin_border
    ws[f"{start_col_letter}4"].border = thin_border
    ws[f"{get_column_letter(start_column+2)}2"].border = thin_border
    ws[f"{get_column_letter(start_column + 1)}3"].border = thin_border
    ws[f"{get_column_letter(start_column + 2)}3"].border = thin_border
    ws[f"{get_column_letter(start_column + 1)}4"].border = thin_border
    ws[f"{get_column_letter(start_column + 2)}4"].border = thin_border
    
    ws[f"{get_column_letter(start_column + 1)}3"].fill = header_fill
    ws[f"{get_column_letter(start_column + 1)}4"].fill = header_fill
    
    row = 6
    contador = 1
    df_periodos = df_agrupado.groupby('periodo')

    # Iterar sobre los periodos agrupados
    for periodo, datos in df_periodos:
        # Ajustar las posiciones de acuerdo a start_column
        horas = calculate_hours(datos["periodo"].iloc[0])
        ws[f"{start_col_letter}{row}"] = f"BLOQUE {contador}"
        ws[f"{get_column_letter(start_column + 1)}{row}"] = "SUBESTACIÓN"
        ws[f"{get_column_letter(start_column + 2)}{row}"] = "PRIMARIOS A DESCONECTAR"
        ws[f"{get_column_letter(start_column + 3)}{row}"] = '# CLIENTES'
        ws[f"{get_column_letter(start_column + 6)}{row}"] = 'DEMANDA PROMEDIO DE LOS PERIODOS (MWh)'
        ws[f"{get_column_letter(start_column + 9)}{row}"] = "PROVINCIA"
        ws[f"{get_column_letter(start_column + 10)}{row}"] = "CANTON"
        ws[f"{get_column_letter(start_column + 11)}{row}"] = "SECTORES"


        ws[f"{get_column_letter(start_column + 3)}{row+1}"] = 'RESIDENCIAL'
        ws[f"{get_column_letter(start_column + 4)}{row+1}"] = 'INDUSTRIAL'
        ws[f"{get_column_letter(start_column + 5)}{row+1}"] = 'COMERCIAL'
        ws[f"{get_column_letter(start_column + 6)}{row+1}"] = 'RESIDENCIAL'
        ws[f"{get_column_letter(start_column + 7)}{row+1}"] = 'INDUSTRIAL'
        ws[f"{get_column_letter(start_column + 8)}{row+1}"] = 'COMERCIAL'

        # Aplicar formato a las celdas del encabezado
# Aplicar estilos a la fila actual y la siguiente fila en un solo bucle
        for col in range(start_column, start_column + 12):
            for r in range(row, row + 2):  # Aplica los estilos a la fila actual (row) y la siguiente (row + 1)
                cell = ws.cell(row=r, column=col)
                cell.font = bold_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws.merge_cells(f"{start_col_letter}{row}:{start_col_letter}{row+1}")
        ws.merge_cells(f"{get_column_letter(start_column + 1)}{row}:{get_column_letter(start_column + 1)}{row+1}")
        ws.merge_cells(f"{get_column_letter(start_column + 2)}{row}:{get_column_letter(start_column + 2)}{row+1}")        
        ws.merge_cells(f"{get_column_letter(start_column + 9)}{row}:{get_column_letter(start_column + 9)}{row+1}")        
        ws.merge_cells(f"{get_column_letter(start_column + 10)}{row}:{get_column_letter(start_column + 10)}{row+1}")        
        ws.merge_cells(f"{get_column_letter(start_column + 11)}{row}:{get_column_letter(start_column + 11)}{row+1}")        

        ws.merge_cells(f"{get_column_letter(start_column + 3)}{row}:{get_column_letter(start_column + 5)}{row}")        
        ws.merge_cells(f"{get_column_letter(start_column + 6)}{row}:{get_column_letter(start_column + 8)}{row}")        

        row +=1

        if isinstance(datos, pd.Series):
            datos = datos.to_frame().T
        merge_start = row + 1
        for _, fila in datos.iterrows():
            carga_est = float(fila["carga_est_mw"])
            # Insertar fila a partir de start_column
            ws.append([None] * (start_column - 1) + list(fila[:-1]))
            # Recuperar la fila actual donde estamos añadiendo datos
            current_row = ws.max_row
            # Recuperar el valor actual en la columna start_column + 6
            current_value = ws[f"{get_column_letter(start_column + 6)}{current_row}"].value
            ws[f"{get_column_letter(start_column + 6)}{current_row}"] = f"={current_value}*{horas}*{carga_est}"
            current_value = ws[f"{get_column_letter(start_column + 7)}{current_row}"].value
            ws[f"{get_column_letter(start_column + 7)}{current_row}"] = f"={current_value}*{horas}*{carga_est}"
            current_value = ws[f"{get_column_letter(start_column + 8)}{current_row}"].value
            ws[f"{get_column_letter(start_column + 8)}{current_row}"] = f"={current_value}*{horas}*{carga_est}"

            row += 1
        merge_end = row

        # Aplicar borde a las celdas del rango de datos
        for r in range(row - len(datos), row + 3):
            for c in range(start_column, start_column + len(fila[:-1])):
                cell = ws.cell(row=r, column=c)
                cell.border = thin_border

        contador += 1
        row += 4
        #MErge periodos
        ws.merge_cells(f'{start_col_letter}{merge_start}:{start_col_letter}{merge_end+2}')

        # Centrar el contenido después del merge
        merged_cell = ws[f"{start_col_letter}{merge_start}"]
        merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)



        # Usar la fórmula de Excel para sumar el rango
        ws[f"{get_column_letter(start_column + 1)}{merge_end + 1}"] = "TOTALES PARCIALES:"
        ws[f"{get_column_letter(start_column + 1)}{merge_end + 2}"] = "TOTAL:"

        
        ws[f"{get_column_letter(start_column + 3)}{merge_end + 1}"] = f"=SUM({get_column_letter(start_column + 3)}{merge_start}:{get_column_letter(start_column + 3)}{merge_end})"
        ws[f"{get_column_letter(start_column + 4)}{merge_end + 1}"] = f"=SUM({get_column_letter(start_column + 4)}{merge_start}:{get_column_letter(start_column + 4)}{merge_end})"
        ws[f"{get_column_letter(start_column + 5)}{merge_end + 1}"] = f"=SUM({get_column_letter(start_column + 5)}{merge_start}:{get_column_letter(start_column + 5)}{merge_end})"
        ws[f"{get_column_letter(start_column + 6)}{merge_end + 1}"] = f"=SUM({get_column_letter(start_column + 6)}{merge_start}:{get_column_letter(start_column + 6)}{merge_end})"
        ws[f"{get_column_letter(start_column + 7)}{merge_end + 1}"] = f"=SUM({get_column_letter(start_column + 7)}{merge_start}:{get_column_letter(start_column + 7)}{merge_end})"
        ws[f"{get_column_letter(start_column + 8)}{merge_end + 1}"] = f"=SUM({get_column_letter(start_column + 8)}{merge_start}:{get_column_letter(start_column + 8)}{merge_end})"

        # Sumar las tres primeras columnas (start_column + 3, start_column + 4, start_column + 5) usando SUM
        ws[f"{get_column_letter(start_column + 3)}{merge_end + 2}"] = (
            f"=SUM({get_column_letter(start_column + 3)}{merge_end + 1}, " 
            f"{get_column_letter(start_column + 4)}{merge_end + 1}, "
            f"{get_column_letter(start_column + 5)}{merge_end + 1})"
        )

        # Sumar las siguientes tres columnas (start_column + 6, start_column + 7, start_column + 8) usando SUM
        ws[f"{get_column_letter(start_column + 6)}{merge_end + 2}"] = (
            f"=SUM({get_column_letter(start_column + 6)}{merge_end + 1}, "
            f"{get_column_letter(start_column + 7)}{merge_end + 1}, "
            f"{get_column_letter(start_column + 8)}{merge_end + 1})"
        )


        #Merges para calculos
        ws.merge_cells(f"{get_column_letter(start_column + 1)}{merge_end + 1}:{get_column_letter(start_column + 2)}{merge_end + 1}")
        ws.merge_cells(f"{get_column_letter(start_column + 1)}{merge_end + 2}:{get_column_letter(start_column + 2)}{merge_end + 2}")

        ws.merge_cells(f"{get_column_letter(start_column + 3)}{merge_end + 2}:{get_column_letter(start_column + 5)}{merge_end + 2}")
        ws.merge_cells(f"{get_column_letter(start_column + 6)}{merge_end + 2}:{get_column_letter(start_column + 8)}{merge_end + 2}")
        
        ws.merge_cells(f"{get_column_letter(start_column + 9)}{merge_end + 1}:{get_column_letter(start_column + 11)}{merge_end + 2}")

        #Formato
        ws[f"{get_column_letter(start_column + 1)}{merge_end + 1}"].font = bold_font
        ws[f"{get_column_letter(start_column + 1)}{merge_end + 2}"].font = bold_font
        
        for r in range(merge_start, merge_end+3):
            ws.cell(row=r, column=start_column + 3).number_format = '0.0'
            ws.cell(row=r, column=start_column + 4).number_format = '0.0'
            ws.cell(row=r, column=start_column + 5).number_format = '0.0'
            ws.cell(row=r, column=start_column + 6).number_format = '0.00'
            ws.cell(row=r, column=start_column + 7).number_format = '0.00'
            ws.cell(row=r, column=start_column + 8).number_format = '0.00'
            ws.cell(row=r, column=start_column + 11).alignment = Alignment(vertical='center', wrap_text=True)
            # Alinear todas las columnas a la izquierda
            ws.cell(row=r, column=start_column + 1).alignment = Alignment(horizontal='left',vertical='top')
            ws.cell(row=r, column=start_column + 3).alignment = Alignment(horizontal='left',vertical='top')
            ws.cell(row=r, column=start_column + 4).alignment = Alignment(horizontal='left',vertical='top')
            ws.cell(row=r, column=start_column + 5).alignment = Alignment(horizontal='left',vertical='top')
            ws.cell(row=r, column=start_column + 6).alignment = Alignment(horizontal='left',vertical='top')
            ws.cell(row=r, column=start_column + 7).alignment = Alignment(horizontal='left',vertical='top')
            ws.cell(row=r, column=start_column + 8).alignment = Alignment(horizontal='left',vertical='top')

            ws.cell(row=r, column=start_column + 2).alignment = Alignment(vertical='top')
            ws.cell(row=r, column=start_column + 9).alignment = Alignment(vertical='top')
            ws.cell(row=r, column=start_column + 10).alignment = Alignment(vertical='top')


    # Ajustar ancho de las columnas dinámicamente a partir de start_column
    ancho_columnas = {
        0: 20,  # Columna A (20)
        1: 16,  # Columna B (16)
        2: 25,  # Columna C (25)
        3: 15,  # Columna D (15)
        4: 15,  # Columna E (15)
        5: 15,  # Columna F (15)
        6: 15,  # Columna G (20)
        7: 15,  # Columna H (15)
        8: 15,   # Columna I (15)
        9: 15,   # Columna J (15)
        10: 15,   # Columna K (15)
        11: 180,   # Columna L (15)
    }

    for idx, ancho in ancho_columnas.items():
        col_letter = get_column_letter(start_column + idx)
        ws.column_dimensions[col_letter].width = ancho